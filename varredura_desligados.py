"""
varredura_desligados.py
=======================
Lê os CSVs de funcionários desligados da Rede Ápice, cruza com
usuários ativos no Activesoft (6 unidades) e no MS365 (Azure AD),
desativa quem ainda está ativo e envia relatório Excel por e-mail.

Uso:
    python varredura_desligados.py                        # lê todos os .csv da pasta
    python varredura_desligados.py arq1.csv arq2.csv ...  # arquivos específicos
"""

import os
import sys
import glob
import re
import smtplib
import requests
import msal
import pandas as pd
from datetime import datetime
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email import encoders
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

load_dotenv()

# ── MS365 ──────────────────────────────────────────────────────────────────
MS_TENANT_ID     = os.getenv("MS_TENANT_ID")
MS_CLIENT_ID     = os.getenv("MS_CLIENT_ID")
MS_CLIENT_SECRET = os.getenv("MS_CLIENT_SECRET")
MS_GRAPH_URL     = "https://graph.microsoft.com/v1.0"

# ── Activesoft ─────────────────────────────────────────────────────────────
ACTIVESOFT_BASE_URL = "https://siga.activesoft.com.br"
TOKENS = {
    "OVM":          os.getenv("ACTIVESOFT_TOKEN_OVM"),
    "APS":          os.getenv("ACTIVESOFT_TOKEN_APS"),
    "MARILLAC":     os.getenv("ACTIVESOFT_TOKEN_MARILLAC"),
    "TUTOR":        os.getenv("ACTIVESOFT_TOKEN_TUTOR"),
    "DOMHENRIQUE":  os.getenv("ACTIVESOFT_TOKEN_DH1"),
    "DOMHENRIQUE2": os.getenv("ACTIVESOFT_TOKEN_DH2"),
}

# ── SMTP ───────────────────────────────────────────────────────────────────
SMTP_HOST     = os.getenv("SMTP_HOST",     "smtp.office365.com")
SMTP_PORT     = int(os.getenv("SMTP_PORT", 587))
SMTP_USER     = os.getenv("SMTP_USER",     "adm.bot@redeapice.com.br")
SMTP_PASSWORD = os.getenv("SMTP_PASSWORD")
EMAIL_DESTINO = os.getenv("EMAIL_DESTINO", "bruno.fernandes@redeapice.com.br")

# ── Mapeamento arquivo → unidade Activesoft ────────────────────────────────
# Ajuste se os nomes dos arquivos mudarem
MAPA_UNIDADE = {
    "TUTOR":           "TUTOR",
    "MARILLAC":        "MARILLAC",
    "DOM_HENRIQUE_II": "DOMHENRIQUE2",
    "DOM_HENRIQUE_I":  "DOMHENRIQUE",
    "APICE_REDE":      "OVM",       # ajuste conforme necessário
    "APICE_PINHEIROS": "APS",
    "VILA_MERCES":     "OVM",       # sem token próprio — ajuste se necessário
    "COLEGIO_VILA":    "OVM",
}

def _unidade_do_arquivo(nome_arquivo: str) -> str:
    nome = nome_arquivo.upper()
    for chave, unidade in MAPA_UNIDADE.items():
        if chave in nome:
            return unidade
    return "TODAS"  # cruzar em todas se não identificar

# ─────────────────────────────────────────────────────────────────────────────
# LEITURA DOS CSVs
# ─────────────────────────────────────────────────────────────────────────────
def normalizar_cpf(valor) -> str:
    if pd.isna(valor):
        return ""
    cpf = re.sub(r"\D", "", str(valor))
    return cpf if len(cpf) == 11 else ""

def ler_csvs(arquivos: list) -> pd.DataFrame:
    frames = []
    for arq in arquivos:
        try:
            df = pd.read_csv(arq, sep=";", dtype=str, encoding="utf-8-sig")
            df.columns = [c.strip().upper() for c in df.columns]

            # CPF
            col_cpf = next((c for c in df.columns if "CPF" in c), None)
            # Nome
            col_nome = next((c for c in df.columns if "NOME" in c), None)
            # Matrícula: REGISTRO_DO_COLABORADOR tem precedência, senão CÓDIGO
            col_mat = next((c for c in df.columns if "REGISTRO" in c), None) or \
                      next((c for c in df.columns if "CÓDIGO" in c or "CODIGO" in c), None)
            # Último dia
            col_data = next((c for c in df.columns if "ÚLTIMO" in c or "ULTIMO" in c), None)

            rows = []
            for _, row in df.iterrows():
                cpf  = normalizar_cpf(row.get(col_cpf, ""))
                mat  = str(row.get(col_mat, "")).strip() if col_mat else ""
                nome = str(row.get(col_nome, "")).strip() if col_nome else ""
                data = str(row.get(col_data, "")).strip() if col_data else ""

                if not cpf and not mat:
                    continue

                rows.append({
                    "nome":         nome,
                    "cpf":          cpf,
                    "matricula":    mat,
                    "ultimo_dia":   data,
                    "unidade_hint": _unidade_do_arquivo(os.path.basename(arq)),
                    "origem":       os.path.basename(arq),
                })

            frames.append(pd.DataFrame(rows))
            print(f"  ✔ {os.path.basename(arq)}: {len(rows)} funcionários")

        except Exception as e:
            print(f"  ✘ Erro em {arq}: {e}")

    if not frames:
        raise SystemExit("Nenhum funcionário encontrado nos arquivos.")

    df_total = pd.concat(frames, ignore_index=True).drop_duplicates(subset=["cpf", "matricula"])
    print(f"\n📋 Total de desligados únicos: {len(df_total)}\n")
    return df_total

# ─────────────────────────────────────────────────────────────────────────────
# MS365
# ─────────────────────────────────────────────────────────────────────────────
def _token_ms365() -> str:
    app = msal.ConfidentialClientApplication(
        MS_CLIENT_ID,
        authority=f"https://login.microsoftonline.com/{MS_TENANT_ID}",
        client_credential=MS_CLIENT_SECRET,
    )
    res = app.acquire_token_for_client(scopes=["https://graph.microsoft.com/.default"])
    if "access_token" not in res:
        raise Exception(f"Falha token MS365: {res.get('error_description')}")
    return res["access_token"]

def _buscar_ms365(token: str, cpf: str = "", matricula: str = "") -> dict | None:
    headers = {"Authorization": f"Bearer {token}"}
    tentativas = []
    if cpf:
        tentativas += [
            f"onPremisesExtensionAttributes/extensionAttribute1 eq '{cpf}'",
            f"onPremisesExtensionAttributes/extensionAttribute2 eq '{cpf}'",
        ]
    if matricula:
        tentativas.append(f"employeeId eq '{matricula}'")

    for filtro in tentativas:
        url = (
            f"{MS_GRAPH_URL}/users"
            f"?$filter={requests.utils.quote(filtro)}"
            f"&$select=id,displayName,userPrincipalName,accountEnabled,employeeId"
        )
        try:
            resp = requests.get(url, headers=headers, timeout=15)
            if resp.status_code == 200:
                dados = resp.json().get("value", [])
                if dados:
                    return dados[0]
        except Exception:
            pass
    return None

def desativar_ms365(token: str, user_id: str) -> bool:
    resp = requests.patch(
        f"{MS_GRAPH_URL}/users/{user_id}",
        headers={"Authorization": f"Bearer {token}", "Content-Type": "application/json"},
        json={"accountEnabled": False},
        timeout=15,
    )
    return resp.status_code in (200, 204)

def cruzar_ms365(df: pd.DataFrame, token: str) -> list:
    resultados = []
    total = len(df)
    for i, (_, row) in enumerate(df.iterrows(), 1):
        print(f"  MS365 [{i}/{total}] {row['nome'][:40]:<40}", end="\r")
        u = _buscar_ms365(token, cpf=row["cpf"], matricula=row["matricula"])
        if u and u.get("accountEnabled"):
            resultados.append({
                "sistema":       "MS365",
                "unidade":       "—",
                "nome_sistema":  u.get("displayName", ""),
                "nome_planilha": row["nome"],
                "login":         u.get("userPrincipalName", ""),
                "cpf":           row["cpf"],
                "matricula":     row["matricula"],
                "ultimo_dia":    row["ultimo_dia"],
                "origem":        row["origem"],
                "user_id":       u.get("id"),
                "status_acao":   None,
            })
    print()
    return resultados

# ─────────────────────────────────────────────────────────────────────────────
# ACTIVESOFT
# ─────────────────────────────────────────────────────────────────────────────
def _headers_as(unidade: str) -> dict:
    token = TOKENS.get(unidade.upper())
    if not token:
        raise Exception(f"Token não configurado: {unidade}")
    return {
        "Authorization": f"Bearer {token}",
        "Accept": "application/json",
        "X-Requested-With": "XMLHttpRequest",
    }

def _listar_as(unidade: str) -> list:
    resp = requests.get(
        f"{ACTIVESOFT_BASE_URL}/api/v0/lista_usuarios/?version=0",
        headers=_headers_as(unidade),
        timeout=15,
    )
    resp.raise_for_status()
    u = resp.json()
    if not isinstance(u, list):
        u = u.get("results") or u.get("data") or []
    return u

def desativar_as(unidade: str, login: str) -> bool:
    try:
        resp = requests.post(
            f"{ACTIVESOFT_BASE_URL}/api/v0/usuario/{login}/desativar/",
            headers=_headers_as(unidade),
            timeout=15,
        )
        if resp.status_code == 404:
            resp = requests.patch(
                f"{ACTIVESOFT_BASE_URL}/api/v0/usuario/{login}/",
                headers={**_headers_as(unidade), "Content-Type": "application/json"},
                json={"ativo": False},
                timeout=15,
            )
        return resp.status_code in (200, 201, 204)
    except Exception:
        return False

import unicodedata as _ud

def _normalizar(texto: str) -> str:
    texto = _ud.normalize("NFD", str(texto).lower())
    texto = "".join(c for c in texto if _ud.category(c) != "Mn")
    return re.sub(r"[^a-z0-9 ]", " ", texto).strip()

def _login_bate_nome(login: str, nome_planilha: str) -> bool:
    """Compara login (nome.sobrenome) com nome completo da planilha."""
    partes_login  = set(_normalizar(login).replace(".", " ").split())
    palavras_nome = set(w for w in _normalizar(nome_planilha).split() if len(w) > 2)
    return len(partes_login & palavras_nome) >= 2

def _montar_achado_as(u: dict, row, unidade: str) -> dict:
    return {
        "sistema":       "Activesoft",
        "unidade":       unidade,
        "nome_sistema":  u.get("nome", ""),
        "nome_planilha": row["nome"],
        "login":         u.get("login", ""),
        "cpf":           row["cpf"],
        "matricula":     row["matricula"],
        "ultimo_dia":    row["ultimo_dia"],
        "origem":        row["origem"],
        "user_id":       None,
        "status_acao":   None,
    }

def cruzar_activesoft(df: pd.DataFrame) -> list:
    resultados = []
    logins_ja_adicionados = set()
    cpfs_desligados = set(df[df["cpf"] != ""]["cpf"])

    for unidade in TOKENS:
        print(f"  Activesoft [{unidade}] — buscando...", end=" ")
        try:
            usuarios = _listar_as(unidade)
        except Exception as e:
            print(f"✘ Erro: {e}")
            continue

        ativos = [u for u in usuarios if u.get("ativo") is True]
        print(f"{len(ativos)} ativos")

        for u in ativos:
            login = u.get("login", "")
            chave = f"{unidade}:{login}"
            if chave in logins_ja_adicionados:
                continue

            # Estratégia 1: login é um CPF
            cpf_login = normalizar_cpf(login)
            if cpf_login and cpf_login in cpfs_desligados:
                match = df[df["cpf"] == cpf_login]
                if not match.empty:
                    resultados.append(_montar_achado_as(u, match.iloc[0], unidade))
                    logins_ja_adicionados.add(chave)
                    continue

            # Estratégia 2: cruzamento por nome via login (nome.sobrenome)
            for _, row in df.iterrows():
                if row["nome"] and _login_bate_nome(login, row["nome"]):
                    resultados.append(_montar_achado_as(u, row, unidade))
                    logins_ja_adicionados.add(chave)
                    break

    return resultados

# ─────────────────────────────────────────────────────────────────────────────
# DESATIVAÇÕES
# ─────────────────────────────────────────────────────────────────────────────
def executar_desativacoes(achados: list, token_ms365: str) -> list:
    print("\n⚙️  Executando desativações...\n")
    for item in achados:
        if item["sistema"] == "MS365":
            ok = desativar_ms365(token_ms365, item["user_id"])
            item["status_acao"] = "✅ Desativado" if ok else "❌ Falha"
            print(f"  MS365         | {item['login']:<40} → {item['status_acao']}")
        elif item["sistema"] == "Activesoft":
            ok = desativar_as(item["unidade"], item["login"])
            item["status_acao"] = "✅ Desativado" if ok else "❌ Falha"
            print(f"  AS [{item['unidade']:<12}] | {item['login']:<30} → {item['status_acao']}")
    return achados

# ─────────────────────────────────────────────────────────────────────────────
# RELATÓRIO EXCEL
# ─────────────────────────────────────────────────────────────────────────────
COR_HEADER   = "1F3864"
COR_VERDE    = "C6EFCE"
COR_VERMELHO = "FFC7CE"
COR_AMARELO  = "FFEB9C"
COR_CINZA    = "F2F2F2"

def _hdr(cell):
    cell.font      = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    cell.fill      = PatternFill("solid", start_color=COR_HEADER)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

def _borda():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def gerar_relatorio(achados: list, df_desligados: pd.DataFrame) -> str:
    ts    = datetime.now().strftime("%Y%m%d_%H%M%S")
    path  = os.path.join(os.getcwd(), f"relatorio_desligados_{ts}.xlsx")
    wb    = Workbook()

    # ── Aba 1: Desativações ───────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Desativações"
    ws1.row_dimensions[1].height = 28
    cabecalhos = ["Sistema","Unidade","Nome (Sistema)","Nome (Planilha)",
                  "Login / UPN","CPF","Matrícula","Último Dia","Arquivo","Status"]
    ws1.append(cabecalhos)
    for cell in ws1[1]:
        _hdr(cell)
    for col, w in zip("ABCDEFGHIJ", [12,14,28,28,36,15,14,13,34,18]):
        ws1.column_dimensions[col].width = w

    for item in achados:
        ws1.append([
            item["sistema"], item["unidade"], item["nome_sistema"],
            item["nome_planilha"], item["login"], item["cpf"],
            item["matricula"], item["ultimo_dia"], item["origem"],
            item.get("status_acao", "—"),
        ])
        ln  = ws1.max_row
        st  = str(item.get("status_acao", ""))
        cor = COR_VERDE if "✅" in st else COR_VERMELHO if "❌" in st else COR_AMARELO
        for col in range(1, 11):
            c = ws1.cell(row=ln, column=col)
            c.fill      = PatternFill("solid", start_color=cor)
            c.border    = _borda()
            c.alignment = Alignment(vertical="center")
            c.font      = Font(name="Arial", size=10)

    # ── Aba 2: Não encontrados ─────────────────────────────────────────────
    ws2 = wb.create_sheet("Não Encontrados")
    ws2.row_dimensions[1].height = 28
    cpfs_enc = {i["cpf"] for i in achados if i["cpf"]}
    mats_enc = {i["matricula"] for i in achados if i["matricula"]}
    nao_enc  = df_desligados[
        ~df_desligados["cpf"].isin(cpfs_enc) &
        ~df_desligados["matricula"].isin(mats_enc)
    ]
    ws2.append(["Nome","CPF","Matrícula","Último Dia","Arquivo"])
    for cell in ws2[1]:
        _hdr(cell)
    for _, r in nao_enc.iterrows():
        ws2.append([r["nome"], r["cpf"], r["matricula"], r["ultimo_dia"], r["origem"]])
        ln = ws2.max_row
        for col in range(1, 6):
            c = ws2.cell(row=ln, column=col)
            c.fill   = PatternFill("solid", start_color=COR_CINZA)
            c.border = _borda()
            c.font   = Font(name="Arial", size=10)
    for col, w in zip("ABCDE", [28,15,14,13,34]):
        ws2.column_dimensions[col].width = w

    # ── Aba 3: Resumo ──────────────────────────────────────────────────────
    ws3 = wb.create_sheet("Resumo")
    ws3.column_dimensions["A"].width = 40
    ws3.column_dimensions["B"].width = 20
    ws3["A1"] = "📊 Resumo da Varredura — Rede Ápice de Ensino"
    ws3["A1"].font = Font(name="Arial", size=13, bold=True, color=COR_HEADER)
    ws3.merge_cells("A1:B1")
    ws3.row_dimensions[1].height = 28

    total_ok    = sum(1 for i in achados if "✅" in str(i.get("status_acao", "")))
    total_falha = sum(1 for i in achados if "❌" in str(i.get("status_acao", "")))

    dados = [
        ("Data/Hora da varredura",             datetime.now().strftime("%d/%m/%Y %H:%M:%S")),
        ("Funcionários nas planilhas",          len(df_desligados)),
        ("Encontrados ainda ativos",            len(achados)),
        ("  → MS365",                           sum(1 for i in achados if i["sistema"]=="MS365")),
        ("  → Activesoft",                      sum(1 for i in achados if i["sistema"]=="Activesoft")),
        ("Desativações bem-sucedidas ✅",        total_ok),
        ("Falhas na desativação ❌",             total_falha),
        ("Não encontrados em nenhum sistema",   len(nao_enc)),
    ]
    for i, (label, valor) in enumerate(dados, 2):
        ws3.cell(row=i, column=1, value=label).font  = Font(name="Arial", size=11, bold=True)
        c = ws3.cell(row=i, column=2, value=valor)
        c.font      = Font(name="Arial", size=11)
        c.alignment = Alignment(horizontal="center")
        c.fill      = PatternFill("solid", start_color=COR_CINZA)
        c.border    = _borda()

    wb.save(path)
    print(f"\n📄 Relatório salvo em: {path}")
    return path

# ─────────────────────────────────────────────────────────────────────────────
# EMAIL
# ─────────────────────────────────────────────────────────────────────────────
def enviar_email(path_relatorio: str, achados: list, df_desligados: pd.DataFrame):
    total_enc = len(achados)
    total_ok  = sum(1 for i in achados if "✅" in str(i.get("status_acao", "")))
    total_nok = sum(1 for i in achados if "❌" in str(i.get("status_acao", "")))

    msg = MIMEMultipart()
    msg["From"]    = SMTP_USER
    msg["To"]      = EMAIL_DESTINO
    msg["Subject"] = f"[TI Ápice] Varredura de Desligados — {datetime.now().strftime('%d/%m/%Y %H:%M')}"

    corpo = f"""
<html><body style="font-family:Arial,sans-serif;font-size:14px;">
<h2 style="color:#1F3864;">📋 Varredura de Funcionários Desligados</h2>
<p>Varredura concluída em <strong>{datetime.now().strftime('%d/%m/%Y às %H:%M')}</strong>.</p>

<table border="1" cellpadding="6" cellspacing="0" style="border-collapse:collapse;width:420px;">
  <tr style="background:#1F3864;color:#fff;"><td colspan="2"><strong>Resumo</strong></td></tr>
  <tr><td>Funcionários nas planilhas</td><td><strong>{len(df_desligados)}</strong></td></tr>
  <tr><td>Encontrados ainda ativos</td><td><strong>{total_enc}</strong></td></tr>
  <tr><td>Desativações bem-sucedidas</td><td style="color:green;"><strong>{total_ok} ✅</strong></td></tr>
  <tr><td>Falhas na desativação</td><td style="color:red;"><strong>{total_nok} ❌</strong></td></tr>
</table>

<p>O relatório detalhado está em anexo.</p>
<p style="color:#888;font-size:12px;">Enviado automaticamente pelo bot TI — Rede Ápice de Ensino</p>
</body></html>
"""
    msg.attach(MIMEText(corpo, "html"))

    with open(path_relatorio, "rb") as f:
        parte = MIMEBase("application", "octet-stream")
        parte.set_payload(f.read())
        encoders.encode_base64(parte)
        parte.add_header(
            "Content-Disposition",
            f"attachment; filename={os.path.basename(path_relatorio)}"
        )
        msg.attach(parte)

    try:
        with smtplib.SMTP(SMTP_HOST, SMTP_PORT) as smtp:
            smtp.ehlo()
            smtp.starttls()
            smtp.login(SMTP_USER, SMTP_PASSWORD)
            smtp.sendmail(SMTP_USER, EMAIL_DESTINO, msg.as_string())
        print(f"📧 E-mail enviado para {EMAIL_DESTINO}")
    except Exception as e:
        print(f"⚠️  Falha ao enviar e-mail: {e}")

# ─────────────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────────────
def main():
    print("=" * 62)
    print("  VARREDURA DE FUNCIONÁRIOS DESLIGADOS — Rede Ápice de Ensino")
    print(f"  {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}")
    print("=" * 62)

    arquivos = sys.argv[1:] or sorted(glob.glob("*.csv"))
    arquivos = [a for a in arquivos if not a.startswith("relatorio_")]
    if not arquivos:
        raise SystemExit("❌ Nenhum arquivo .csv encontrado. Passe como argumento ou coloque na mesma pasta.")

    print(f"\n📂 Arquivos: {[os.path.basename(a) for a in arquivos]}\n")

    # 1. Carrega desligados
    df = ler_csvs(arquivos)

    # 2. Token MS365
    print("🔑 Obtendo token MS365...")
    token_ms365 = _token_ms365()
    print("   ✔ OK\n")

    # 3. Cruza Activesoft
    print("🔎 Cruzando com Activesoft...\n")
    achados_as = cruzar_activesoft(df)
    print(f"\n   → {len(achados_as)} encontrado(s) no Activesoft\n")

    # 4. Cruza MS365
    print("🔎 Cruzando com MS365...\n")
    achados_ms = cruzar_ms365(df, token_ms365)
    print(f"   → {len(achados_ms)} encontrado(s) no MS365\n")

    achados = achados_as + achados_ms

    if not achados:
        print("✅ Nenhum funcionário desligado encontrado ativo nos sistemas.\n")
        for item in achados:
            item["status_acao"] = "✅ Não encontrado (ok)"
    else:
        print(f"⚠️  {len(achados)} funcionário(s) ainda ativo(s).\n")
        resp = input("Deseja executar as desativações agora? [s/N] ").strip().lower()
        if resp == "s":
            achados = executar_desativacoes(achados, token_ms365)
        else:
            print("  ℹ️  Desativações puladas.")
            for item in achados:
                item["status_acao"] = "⏸ Não executado"

    # 5. Relatório
    path = gerar_relatorio(achados, df)

    # 6. E-mail
    enviar_email(path, achados, df)


if __name__ == "__main__":
    main()
